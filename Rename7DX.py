import os
import pdftotext
import openpyxl
import ExtractKeyWords
import Tools
import Rename9DW
def rename7DX(filespath,offset):
    Rename9DW.rename("2021026SUSAR-",filespath,offset)

def rename7DX(FilePath):
    # I want the BMS_FU be the key, and ???
    filePath=FilePath
    def f(nums):
        ans=[]
        for i in range(len(nums)):
            if nums[i][:4]=="BMS-" and len(nums[i])==15:
                ans.append(nums[i])
                break    
                
        for i in range(len(nums)):
            if nums[i]=="FOLLOWUP:":
                if len(str(nums[i+1]))<3:
                    ans.append(nums[i+1])
                else:
                    ans.append("0")

        return ans 


    #返回随访类型中的数字，首次为0
    def getnumber(s):
        if not s:
            return "0"
        ret=""
        for i in range(len(s)):
            e=s[i]
            if e.isdigit():
                ret=ret+e
                i+=1
                while i<len(s) and s[i].isdigit():
                    ret+=s[i]
                    i+=1
                return ret
                
        return "0"

    #对于给定的关键字数组 "Ks"，返回对应的列
    def getcolumn(Ks):
        for e in Row:
            tk=e+"1"
            cell=sheet[tk]
            if not cell:
                continue
            ts=cell.value 
            if not ts:
                continue
            flag=0
            for ks in Ks:
                if ks not in ts:
                    flag+=1
            if flag==0:
                return e
    def findRepeat(s):
        return "文件" in s
        """
        if len(s)>3:
            for i in range(len(s)-1):
                if s[i:i+2]=="文件":
                    return 1
        return 0
        """


    #fileolders=get_filelist(FilePath,[])
    def get_filelist(Dir,Filelist):
        newDir=Dir
        if os.path.isfile(Dir):
            Filelist.append(Dir)
        elif os.path.isdir(Dir):
            for s in os.listdir(Dir):
                newDir=os.path.join(Dir,s)
                get_filelist(newDir,Filelist)
        return Filelist
    fileList=get_filelist(FilePath,[])

    Sheets=[]
    for file in fileList:
        if len(file)>5 and (file[len(file)-1:len(file)-6:-1]=='xslx.'):
            print(file)
            wb=openpyxl.load_workbook(file)
            for sheetname in wb.sheetnames:
                Sheets.append(wb[sheetname])
        if len(file)>5 and file[len(file)-1:len(file)-6:-1]=='mslx.':
            wb=openpyxl.load_workbook(file,keep_vba=True)
            for sheetname in wb.sheetnames:
                Sheets.append(wb[sheetname])
    #sheet=wb[wb.sheetnames[0]]
    Row=["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
    K0=["序号"]
    K1=["SUSAR","编号","No"]
    K2=["随访类型"]
    K3=["方案编号"]
    K9=["北肿"]
    KReport=["报告时间"]
    KManufacturer=["获知时间"]
    KSTART=["发生时间"]
    mp={}
    mpReportTime={}
    mpManufacturerTime={}
    mpStart={}
    d={}
    Re=set()
    Proof={}
    date={"JAN":"01","FEB":"02","MAR":"03","APR":"04","MAY":"05","JUN":"06","JUL":"07","AUG":"08","SEP":"09","OCT":"10","NOV":"11","DEC":"12","NPV":"11"}
    def dateconvert(s):

        ret=""
        ret+=s[len(s)-4:]+"-"
        
        ret+=date[s[3:6].upper()]+"-"
        ret+=s[:2]
        ret+="00:00:00"
        return ret
    for sheet in Sheets:
        rows=sheet.max_row  
        cols=sheet.max_column
        k0=getcolumn(K0)    
        k1=getcolumn(K1)
        k2=getcolumn(K2)
        k3=getcolumn(K3)
        k9=getcolumn(K9)
        for i in range(2,rows+1):
            if k1 and k2 and k0 and k3:

                    
                k=str(sheet[k1+str(i)].value).upper()+str(getnumber(sheet[k2+str(i)].value))
                bk=k.replace(" ","")
                k=bk+""
                #key 加上序号
                k=k+str(sheet[k0+str(i)].value)
                if str(sheet[k3+str(i)].value).replace(" ","")=="CA209-7DX":
                    d[k]=1            
                #第一列需要放重命名后的序号
                c1=sheet[k9+str(i)]
                if not c1:
                    continue
                v=str(c1.value)
                if not v.isdigit():
                    continue
                """
                if v in Re:
                    continue
                Re.add(v)
                """
                mp.setdefault(k,[]).append(v)
       
    for file in fileList:
        if len(file)>4 and file[len(file)-1:len(file)-5:-1]=='fdp.':
            #print(file)
            fp=file
            tf=open(fp,"rb")
            pdf=pdftotext.PDF(tf)
            col=f([x for x in pdf[1].split()])
            oldname=fp
            for i in range(len(fp)-1,0,-1):
                if fp[i]=="/":
                    filePath=fp[:i]
                    file=fp[i+1:]
                    break
            if len(col)<2:
                newname=filePath+'/'+"请检查该文件内容！"+file
            elif len(col)==2:
                tn=col[0]+col[1]
                bn=tn.replace(" ","")
                k=bn+""
                #key 后加上序号，要求文件名如"数字+"_"+"文件名""
                
                for i in range(len(file)):
                    if file[i]=="_":
                        k=k+str(file[:i])
                        break
                
                if k not in mp:
                    print(col,k)
                    newname=filePath+'/'+"该文件没在excel中记录！"+file
                    if findRepeat(file):
                        newname=oldname
                else:
                    if k in d:
                        newname=filePath + os.sep+"2021026SUSAR-"+"_可能重复！_".join(mp[k])+".pdf"
                    else:
                        newname=filePath + os.sep+"2021026SUSAR-"+"_可能重复！_".join(mp[k])+"-W.pdf"

            os.rename(oldname,newname)
    print(mp)