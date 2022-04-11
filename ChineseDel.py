#!/usr/bin/python3
from openpyxl import Workbook
import pdftotext
import os
import openpyxl 
import re
import sys 
import logging

#logging.basicConfig(level=logging.DEBUG)
#filePath=sys.argv[1:][0]
#filePath='/mnt/'+filePath[0].lower()+filePath[2:]
#FilePath='/mnt/c/Users/13651/桌面/Delete'
#FilePath=sys.argv[1:][0]
ColumnsNeededToDelete=["Indication 适应症","Gender 性别","Country 国家","SAE 名称",
    "Type Report 随访类型","Severity 严重程度","Death 死亡","Action 措施","Outcome 结局","Ipi 相关性","Nivo 相关性"]
ColumsKeep=["No 序号","Protocol 方案编号","SUSAR 编号","Subject 编号","Age 年龄",
    "Start 发送时间","Manufacturer 获知时间","Report 报告时间"]
def containChinese(s):
    for e in s:
        if '\u4e00' <= e <= '\u9fa5':
            return True
    return False 

def needed(cell):
    # the first row
    if not cell.value:
        return False
    s=str(cell.value)
    for nd in ColumnsNeededToDelete:
        flag=1
        for e in nd.split():
            if e not in s:
                flag=0
                break
        if flag:
            return True
    return False

def keepChinese(value):
    #print(value)
    #logging.debug("textdebug",value)
    # the in put is chinese
    if not isinstance(value,str):
        return value
    ret=[]
    for e in value.split('\n'):
        if containChinese(e):
            ret.append(e)
    return '\n'.join(ret)

def chineseDel(FilePath):

    for fp in os.walk(FilePath):
        for filename in fp[2]:
            if os.path.splitext(filename)[1]==".xlsm" or os.path.splitext(filename)[1]==".xlsx":
                #print(os.path.splitext(filename)[1],os.path.splitext(filename)[1]==".xlsm")
                if os.path.splitext(filename)[1]==".xlsm":
                    wb=openpyxl.load_workbook(fp[0]+os.sep+filename,keep_vba=True)
                else:
                    wb=openpyxl.load_workbook(fp[0]+os.sep+filename)
            for sheetname in wb.sheetnames:
                ws=wb[sheetname]
                #预过滤sheet
                #print(ws.max_column)
                #logging.debug("ws.max_column is %s",ws.max_column)
                if ws.max_column<20:
                    continue
                for j in range(1,ws.max_column+1):
                    if not needed(ws.cell(1,j)):
                        #logging.debug("test cell need or not %s",ws.cell(1,j).value)
                        continue
                    for i in range(2,ws.max_row+1):
                        cell=ws.cell(i,j)
                        if not cell.value:
                            continue
                        cell.value=keepChinese(cell.value)                
            wb.save(fp[0]+os.sep+filename)
            wb.close()  

"""
    def getnumber(s):
        if not s:
            return "0"
        ret=""
        for i in range(len(s)):
            e=s[i]
            if e.isdigit():
                ret=ret+e
                i+=1
                while i<len(s) and s[i].isdigit():
                    ret+=s[i]
                    i+=1
                return ret
                
        return "0"
    def get_filelist(Dir,Filelist):
        newDir=Dir
        if os.path.isfile(Dir):
            Filelist.append(Dir)
        elif os.path.isdir(Dir):
            for s in os.listdir(Dir):
                newDir=os.path.join(Dir,s)
                get_filelist(newDir,Filelist)
        return Filelist

    def Exclu(title):
        if not title:
            return 1
        return (code in title) or (code1 in title)
    def Follow(title):
        return code2 in title
    def findCol(s):
        if not s:
            return False
        for key in Needs:
            if key in s:
                return True
        return False


    fileList=get_filelist(FilePath,[])
    Row=["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

    #需要处理的列的关键词
    Needs=['性别','国家','名称','随访','严重程度','事件结局','相关性','采取措施','致死']
    code='号'
    code1='间'
    code2='随'
              






    for fp in fileList:
        if len(fp)>5 and (fp[len(fp)-1:len(fp)-6:-1]=='mslx.' or fp[len(fp)-1:len(fp)-6:-1]=='xslx.'):
            print(fp)
            if fp[len(fp)-1:len(fp)-6:-1]=='mslx.':
                #print('xlsm')
                wb=openpyxl.load_workbook(fp,keep_vba=True)
            else:
                wb=openpyxl.load_workbook(fp)
                
            for sheetname in wb.sheetnames:
                ws=wb[sheetname]
                #预过滤sheet
                if ws.max_column<15:
                    continue
                columns=ws.max_row
                for row in Row:
                    if not ws[row+'1']:
                        continue
                    #检测该列是否需要去英文
                    if not findCol(str(ws[row+'1'].value)):
                        continue
                    for column in range(2,columns+1):
                        #提取文本
                        cell=ws[row+str(column)]
                        if not cell:
                            continue
                        v=cell.value
                        if not v:
                            continue
                        v=str(v)
                        #v=re.sub("[A-Za-z]", "", v)
                        v=re.sub("[^(\u4e00-\u9fa5)(\n)]", "", v)
                        v=v.replace(' ','')
                        if "\n\n" in v:
                            v=v.replace('\n\n','\n')
                            v=v.replace('0','')
                            #print(v)
                        v=v.replace('0','')
                        v=v.replace('_','')
                        v=v.replace(';','\n')
                        v=v.replace('；','\n')
                        if not v:
                            continue
                        if v[0]=='\n':
                            v=v[1:]
                        if Follow(str(ws[row+'1'].value)):
                            number=getnumber(str(cell.value))
                            if number!="0":
                                v=v.replace(number,'')
                                v=v.replace(':','')
                                v=v+'-'+number
                        cell.value=v
                
            wb.save(fp)
            wb.close()
"""