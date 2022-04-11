#!/usr/bin/python3
from distutils.fancy_getopt import wrap_text
from fileinput import filename
import openpyxl
import pdftotext
import sys
import os
import json
import Tools
import xlwt

Cols = ["序号","Protocol方案编号","Indication适应症","SUSAR No. 编号","Subject No.受试者编号",
"Gender性别","Age年龄","Country国家","SAE TermSAE名称","Start Date发生时间",
"ManufacturerReceipt Date获知时间","Report Date报告时间","Type of Safety Report随访类型","Severity严重程度","Death是否致死亡",
"Action taked采取措施","Outcome事件结局","Len Causality(Investigator)Nivo 相关性(研究者)","Len Causality(BMS)Nivo 相关性(BMS)"]
Mounths=["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]

#ProtocalMap={"1":"1","20975":"CA209-7JH","INCB":"CA209-9UJ","NU18C02":"CA017-075","PCI-32765LYM1002":"CA209-330"}
CurrentDir=os.path.dirname(__file__)+os.sep+"static"+os.sep

ProtocalMap = Tools.fileToDict(CurrentDir+"ProtocalMap.txt")
   
CountryDict = Tools.fileToDict(CurrentDir+"Country.txt")

Country=set()
for key in CountryDict:
    for k in key.split():
        Country.add(k)

Indication = Tools.fileToDict(CurrentDir+"Indication.txt")   

def getProtocol(filename,pdf):
    ret=""
    nums=filename.split("_")
    for i in range(len(nums)):
        if nums[i].upper()[:4]=="BMS-" and len(nums[i])==15 and i+3<len(nums):
            ret=nums[i+3].upper().split()[0]
    nums=pdf[0].split()
    # maybe I should just return ret!!!
    if ret:
        return ret
    for i in range(len(nums)):
        if nums[i].upper()[:9]=='PROTOCOL:':
            if len(nums[i])>15:
                return nums[i][9:]# if ret==nums[i][9:].upper() else ret+" --- "+nums[i][9:]
            return nums[i+1]# if ret==nums[i+1].upper() else ret+" --- "+nums[i+1]
    return "1"
def getIndication(key):
    key=key.split()[0]
    #查找适应症 
    return Indication.setdefault(key,"1")    
def getSubject(nums):
    for i in range(len(nums)-3):
        if nums[i]=="Subject" or nums[i]=="Patient":
            if nums[i+1].upper()[:2]=="ID":
                if len(nums[i+1])==2:
                    return nums[i+3]
                elif len(nums[i+1])==3:
                    return nums[i+2]
    return "1"
def getSUSAR(filename,nums):
    for i in range(len(nums)):
        if nums[i][:4].upper()=="BMS-" and len(nums[i])>=15:
            return nums[i]
    nums=filename.split('_') 
    for i in range(len(nums)):
        if nums[i][:4].upper()=="BMS-" and len(nums[i])==15:
            return nums[i]    
    return "1"
def getGender(nums):
    for i in range(len(nums)):
        gender=nums[i].upper()
        if gender=="MALE":
            return "M\n男"
        if gender=="FEMALE":
            return "F\n女"
    return "1"
def getAge(text):
    nums=text.split()
    for i in range(len(nums)):
        gender=nums[i].upper()
        if gender=="MALE":
            if nums[i-1].isdigit() and 1<int(nums[i-1])<120:
                return nums[i-1]
            break
        if gender=="FEMALE":
            if nums[i-1].isdigit() and 1<int(nums[i-1])<120:
                return nums[i-1]
            break
  
    nums=text.split('\n')
    if len(nums)<20:
        return "1"
    for i in range(len(nums)-6):
        for j in range(len(nums[i])):
            if nums[i][j:j+3].upper()=="AGE":
                for r in range(i+1,i+6):
                    tnums=nums[r].split()
                    for k in range(len(tnums)):
                        if tnums[k].isdigit():
                            if tnums[k+1].upper() in Mounths:
                                return "1"
                            return tnums[k]

                return "1"
    return "1"
def getCountry(nums): 
    for i in range(len(nums)):
        if nums[i] in Country:
            for j in range(1,4):
                key=" ".join(nums[i:i+j])
                if key in CountryDict:
                    ret=key+CountryDict[key]
                    for i in range(len(ret)):
                        if '\u4e00' <= ret[i] <= '\u9fa5':
                            return ret[:i]+"\n"+ret[i:]
                    return ret
            #return nums[i]+CountryDict[nums[i]]
    return "1"
def getStartDate(nums):
    for i in range(len(nums)-3):
        if nums[i].isdigit() and nums[i+2].isdigit() and nums[i+1].upper()[:3] in Mounths:
            return "-".join(nums[i:i+3])
    return "1"
def getManufacturerReceiptDate(nums):
    for i in range(len(nums)):
        if nums[i]=='HEALTH':
            return nums[i-1]
    return "1"
def getReportDate(nums):
    for i in range(len(nums)):
        if nums[i]=='INITIAL':
            return nums[i-1]  
    return "1"
def getFollowUp(filename,nums):
    tnums=filename.split('_')
    for i in range(len(tnums)):
        if tnums[i].upper()[:4]=="BMS-" and len(tnums[i])>=15:
            ret=Tools.getDigit(tnums[i+1])
            if ret!=-1:
                return ret
            if tnums[i+1].upper()=="INITIAL":
                return 0
        

    for i in range(len(nums)-2):
        if nums[i].upper()=="FOLLOWUP:":
            if nums[i+1].isdigit() and len(nums[i+1])<3:
                return int(nums[i+1])
            else:
                return 0


    return "1"
def getColumn(file,pdf):
    # pdf is a pdf class, I will need to return the colums
    filename=os.path.split(file)[1]
    page=pdf[1].split()
    text=pdf[1]
    #if len(page)<500:
    #    text=imagePdf2text(file,1)
    #    page=text.split()


    ret=["1"]*21
    #getProtocol
    key=getProtocol(filename,pdf)
    if key in ProtocalMap:
        ret[1]=ProtocalMap[key]
    else:
        ret[1]=key
    #getIndication
    ret[2]=getIndication(ret[1])
    #getSUSAR No.
    ret[3]=getSUSAR(filename,pdf[0].split()+page)
    #get subject No.
    ret[4]=getSubject(page)
    #getGender
    ret[5]=getGender(page)
    #getAge
    ret[6]=getAge(text)
    #get Country
    ret[7]=getCountry(page)
    #get SAE 8
    #get Start Date 9
    ret[9]=getStartDate(page)
    #get Manufactrue date 10
    ret[10]=getManufacturerReceiptDate(page)
    #get report date 11
    ret[11]=getReportDate(page)
    #get Safety Report 12
    ret[12]=getFollowUp(filename,page)
    if ret[12]!="1":
        if ret[12]>0:
            ret[12]="Follow up "+str(ret[12])+"\n随访- "+str(ret[12])
        else:
            ret[12]="Initial\n首次"
    #---
    return ret

def set_style(name,height,bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    style.alignment.wrap=1
    return style


def write_excel(filePath,Rows):
    wb=Tools.openExcel(CurrentDir+os.sep+"Template.xlsm")
    ws=wb["For Linda"]
    for i in range(len(Rows)):
        for j in range(1,ws.max_column+1):
            #print(ws.max_column)
            cell=ws.cell(i+2,j)
            cell.alignment=openpyxl.styles.Alignment(wrap_text=True)
            cell.value=Rows[i][j-1]

    Filename=filePath+".xlsm"
    print(filePath+".xlsm")
    wb.save(Filename)
    wb.close()
    #print(CountryDict,Indication)

def Write_excel(filePath,Columns):
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('For Ling',cell_overwrite_ok=True)
    row0 = ["序号","Protocol方案编号","Indication适应症","SUSAR No. 编号","Subject No.受试者编号","Gender性别","Age年龄","Country国家","SAE TermSAE名称","Start Date发生时间","ManufacturerReceipt Date获知时间","Report Date报告时间","Type of Safety Report随访类型","Severity严重程度","Death是否致死亡","Action taked采取措施","Outcome事件结局","Len Causality(Investigator)Nivo 相关性(研究者)","Len Causality(BMS)Nivo 相关性(BMS)"]

    #colum0 = ["张三","李四","恋习Python","小明","小红","无名"]
    #写第一行
    default=set_style('Times New Roman',220,True)
    for i in range(0,len(row0)):
        sheet1.write(0,i,row0[i],set_style('Times New Roman',220,True))
    c=1
    #写第一列
    for col in Columns:
        col[0]=int(col[0])
        for i in range(len(col)):
            sheet1.write(c,i,col[i],default)
        c+=1
    Filename=filePath+".xls"
    print(filePath+".xls")
    f.save(Filename)
    #print(CountryDict,Indication)

def extractKeyWords(filePath):
    print(filePath)
    #print(list(os.walk(filePath)))
  
    #CountryDict={"AFGHANISTAN":"阿富汗","ALAND ISLANDS":"奥兰群岛","ALBANIA":"阿尔巴尼亚","ALGERIA":"阿尔及利亚","AMERICAN SAMOA":"美属萨摩亚","ANDORRA":"安道尔","ANGOLA":"安哥拉","ANGUILLA":"安圭拉","ANTIGUA AND BARBUDA":"安提瓜和巴布达","ARGENTINA":"阿根廷","ARMENIA":"亚美尼亚","ARUBA":"阿鲁巴","AUSTRALIA":"澳大利亚","AUSTRIA":"奥地利","AZERBAIJAN":"阿塞拜疆","BANGLADESH":"孟加拉","BAHRAIN":"巴林","BAHAMAS":"巴哈马","BARBADOS":"巴巴多斯","BELARUS":"白俄罗斯","BELGIUM":"比利时","BELIZE":"伯利兹","BENIN":"贝宁","BERMUDA":"百慕大","BHUTAN":"不丹","BOLIVIA":"玻利维亚","BOSNIA AND HERZEGOVINA":"波斯尼亚和黑塞哥维那","BOTSWANA":"博茨瓦纳","BOUVET ISLAND":"布维岛","BRAZIL":"巴西","BRUNEI":"文莱","BULGARIA":"保加利亚","BURKINA FASO":"布基纳法索","BURUNDI":"布隆迪","CAMBODIA":"柬埔寨","CAMEROON":"喀麦隆","CANADA":"加拿大","CAPE VERDE":"佛得角","CENTRAL AFRICAN REPUBLIC":"中非","CHAD":"乍得","CHILE":"智利","CHRISTMAS ISLANDS":"圣诞岛","COCOS (KEELING) ISLANDS":"科科斯（基林）群岛","COLOMBIA":"哥伦比亚","COMOROS":"科摩罗","CONGO (CONGO-KINSHASA)":"刚果（金）","CONGO":"刚果","COOK ISLANDS":"库克群岛","COSTA RICA":"哥斯达黎加","COTE D'IVOIRE":"科特迪瓦","CHINA":"中国","CROATIA":"克罗地亚","CUBA":"古巴","CZECH":"捷克","CYPRUS":"塞浦路斯","DENMARK":"丹麦","DJIBOUTI":"吉布提","DOMINICA":"多米尼加","ECUADOR":"厄瓜多尔","EGYPT":"埃及","EQUATORIAL GUINEA":"赤道几内亚","ERITREA":"厄立特里亚","ESTONIA":"爱沙尼亚","ETHIOPIA":"埃塞俄比亚","FAROE ISLANDS":"法罗群岛","FIJI":"斐济","FINLAND":"芬兰","FRANCE":"法国","METROPOLITANFRANCE":"法国大都会","FRENCH GUIANA":"法属圭亚那","FRENCH POLYNESIA":"法属波利尼西亚","GABON":"加蓬","GAMBIA":"冈比亚","GEORGIA":"格鲁吉亚","GERMANY":"德国","GHANA":"加纳","GIBRALTAR":"直布罗陀","GREECE":"希腊","GRENADA":"格林纳达","GUADELOUPE":"瓜德罗普岛","GUAM":"关岛","GUATEMALA":"危地马拉","GUERNSEY":"根西岛","GUINEA-BISSAU":"几内亚比绍","GUINEA":"几内亚","GUYANA":"圭亚那","HAITI":"海地","HONDURAS":"洪都拉斯","HUNGARY":"匈牙利","ICELAND":"冰岛","INDIA":"印度","INDONESIA":"印度尼西亚","IRAN":"伊朗","IRAQ":"伊拉克","IRELAND":"爱尔兰","ISLE OF MAN":"马恩岛","ISRAEL":"以色列","ITALY":"意大利","JAMAICA":"牙买加","JAPAN":"日本","JERSEY":"泽西岛","JORDAN":"约旦","KAZAKHSTAN":"哈萨克斯坦","KENYA":"肯尼亚","KIRIBATI":"基里巴斯","KOREA (SOUTH)":"韩国","KOREA (NORTH)":"朝鲜","KUWAIT":"科威特","KYRGYZSTAN":"吉尔吉斯斯坦","LAOS":"老挝","LATVIA":"拉脱维亚","LEBANON":"黎巴嫩","LESOTHO":"莱索托","LIBERIA":"利比里亚","LIBYA":"利比亚","LIECHTENSTEIN":"列支敦士登","LITHUANIA":"立陶宛","LUXEMBOURG":"卢森堡","MACEDONIA":"马其顿","MALAWI":"马拉维","MALAYSIA":"马来西亚","MADAGASCAR":"马达加斯加","MALDIVES":"马尔代夫","MALI":"马里","MALTA":"马耳他","MARSHALL ISLANDS":"马绍尔群岛","MARTINIQUE":"马提尼克岛","MAURITANIA":"毛里塔尼亚","MAURITIUS":"毛里求斯","MAYOTTE":"马约特","MEXICO":"墨西哥","MICRONESIA":"密克罗尼西亚","MOLDOVA":"摩尔多瓦","MONACO":"摩纳哥","MONGOLIA":"蒙古","MONTENEGRO":"黑山","MONTSERRAT":"蒙特塞拉特","MOROCCO":"摩洛哥","MOZAMBIQUE":"莫桑比克","MYANMAR":"缅甸","NAMIBIA":"纳米比亚","NAURU":"瑙鲁","NEPAL":"尼泊尔","NETHERLANDS":"荷兰","NEW CALEDONIA":"新喀里多尼亚","NEW ZEALAND":"新西兰","NICARAGUA":"尼加拉瓜","NIGER":"尼日尔","NIGERIA":"尼日利亚","NIUE":"纽埃","NORFOLK ISLAND":"诺福克岛","NORWAY":"挪威","OMAN":"阿曼","PAKISTAN":"巴基斯坦","PALAU":"帕劳","PALESTINE":"巴勒斯坦","PANAMA":"巴拿马","PAPUA NEW GUINEA":"巴布亚新几内亚","PERU":"秘鲁","PHILIPPINES":"菲律宾","PITCAIRN ISLANDS":"皮特凯恩群岛","POLAND":"波兰","PORTUGAL":"葡萄牙","PUERTO RICO":"波多黎各","QATAR":"卡塔尔","REUNION":"留尼汪岛","ROMANIA":"罗马尼亚","RWANDA":"卢旺达","RUSSIAN FEDERATION":"俄罗斯联邦","SAINT HELENA":"圣赫勒拿","SAINT KITTS-NEVIS":"圣基茨和尼维斯","SAINT LUCIA":"圣卢西亚","SAINT VINCENT AND THE GRENADINES":"圣文森特和格林纳丁斯","EL SALVADOR":"萨尔瓦多","SAMOA":"萨摩亚","SAN MARINO":"圣马力诺","SAO TOME AND PRINCIPE":"圣多美和普林西比","SAUDI ARABIA":"沙特阿拉伯","SENEGAL":"塞内加尔","SEYCHELLES":"塞舌尔","SIERRA LEONE":"塞拉利昂","SINGAPORE":"新加坡","SERBIA":"塞尔维亚","SLOVAKIA":"斯洛伐克","SLOVENIA":"斯洛文尼亚","SOLOMON ISLANDS":"所罗门群岛","SOMALIA":"索马里","SOUTH AFRICA":"南非","SPAIN":"西班牙","SRI LANKA":"斯里兰卡","SUDAN":"苏丹","SURINAME":"苏里南","SWAZILAND":"斯威士兰","SWEDEN":"瑞典","SWITZERLAND":"瑞士","SYRIA":"叙利亚","TAJIKISTAN":"塔吉克斯坦","TANZANIA":"坦桑尼亚","THAILAND":"泰国","TRINIDAD AND TOBAGO":"特立尼达和多巴哥","TIMOR-LESTE":"东帝汶","TOGO":"多哥","TOKELAU":"托克劳","TONGA":"汤加","TUNISIA":"突尼斯","TURKEY":"土耳其","TURKMENISTAN":"土库曼斯坦","TUVALU":"图瓦卢","UGANDA":"乌干达","UKRAINE":"乌克兰","UNITED ARAB EMIRATES":"阿拉伯联合酋长国","UNITED KINGDOM":"英国","UNITED STATES":"美国","URUGUAY":"乌拉圭","UZBEKISTAN":"乌兹别克斯坦","VANUATU":"瓦努阿图","VATICAN CITY":"梵蒂冈","VENEZUELA":"委内瑞拉","VIETNAM":"越南","WALLIS AND FUTUNA":"瓦利斯群岛和富图纳群岛","WESTERN SAHARA":"西撒哈拉","YEMEN":"也门","YUGOSLAVIA":"南斯拉夫","ZAMBIA":"赞比亚","ZIMBABWE":"津巴布韦","KOREA, REPUBLIC OF":"韩国","HONG KONG":"中国香港","TAIWAN, PROVINCE OF":"CHINA 中国台湾"}

    Rows=[]

    for file in sorted(list(os.walk(filePath))[0][2]):
            
        if len(file)>4 and file[len(file)-1:len(file)-5:-1]=='fdp.':
            fp=filePath+os.sep+file
            #tf=open(fp,"rb")
            #pdf=pdftotext.PDF(tf)
            pdf=Tools.pdfToText(fp)
                #nu=[x for x in pdf[1].split()]
            #col0=f0([x for x in pdf[0].split()])
            #col1=f1([x for x in pdf[1].split()]+[x for x in pdf[2].split()])
            #col=col0+col1     
            col=getColumn(fp,pdf)   
            col[0]=int(file.split("_")[0])

            Rows.append(col)

    write_excel(filePath,Rows)   

