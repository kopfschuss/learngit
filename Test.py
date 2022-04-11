from distutils import text_file
import sys
from traceback import print_tb
from wsgiref.simple_server import sys_version
import ChineseDel

from numpy import flexible
import Tools
import os 
import pdfplumber
import inspect
from PIL import Image
from pdf2image import convert_from_path
import pytesseract
import openpyxl

file = "/mnt/c/Project/workspace/data/pdf/41_Nivo_SUSAR_BMS-2022-030322_FU2_United States_CA017-078 (ROW) Blinded.pdf"
file2="/mnt/c/Project/workspace/data/pdf/212_nivo_ipi_SUSAR_bms-2019-034558_10_DE_CA209-901_ROW_open label.pdf"
#print(pdfToText(file)[1].split())
Cols = ["序号","Protocol方案编号","Indication适应症","SUSAR No. 编号","Subject No.受试者编号",
"Gender性别","Age年龄","Country国家","SAE TermSAE名称","Start Date发生时间",
"ManufacturerReceipt Date获知时间","Report Date报告时间","Type of Safety Report随访类型","Severity严重程度","Death是否致死亡",
"Action taked采取措施","Outcome事件结局","Len Causality(Investigator)Nivo 相关性(研究者)","Len Causality(BMS)Nivo 相关性(BMS)"]
def pathTest(file):
    print(file)
    # for a file, the only import stuff is file dir and filename
    # and the os.path.split will do
    # os.walk will do too
    print("os.path.split",os.path.split(file))
    print("os.path.splitext(file)",os.path.splitext(file))
    filename=os.path.split(file)[1]
    print(file,filename)
    print(os.path.splitext(filename))
    dir=os.path.dirname(file)
    #for tp in os.walk(dir):
    #    print(tp[0],tp[2])
    #    print("\n")
def testopenpyxl(filename):
    if os.path.splitext(filename)[1]==".xlsm":
        wb=openpyxl.load_workbook(filename,keep_vba=True)
    else:# os.path.splitext(filename)==".xlsx":
        wb=openpyxl.load_workbook(filename)

    for ws in wb:
        print(ws)
    ws=wb["Nivo Ipi_Feb22"]
    #print(ws,ws.max_column,ws.max_row)
    #print(ws["A1"])
    #print(ws.cell(ws.max_row+10,ws.max_column+10).value)
    print(ws.max_row)
    for i in range(244,1000):
        for j in range(1,ws.max_column+1):
            #c=ws.cell(i,j)
            #c=ws.cell(243,j)
            ws.cell(i,j).value=ws.cell(243,j).value

    wb.save(os.path.dirname(filename)+os.sep+"test.xlsm")
    wb.close()
    #print(text,"\n",repr(text))

testopenpyxl("/mnt/c/Project/workspace/data/pdf/212 - 副本.xlsm")

#pathTest(file)
#print(len(Cols),Cols)
"""
path=os.path.dirname(file)
walk=os.walk(path)
for w in walk:
    print(len(w))
    print(w)
"""    
#print(list()[0][2])
"""
pdf=pdfToText(file)
for page in pdf:
    print(len(page))
print(pdf[1])
doc = convert_from_path(file)
path, fileName = os.path.split(file)
fileBaseName, fileExtension = os.path.splitext(fileName)


for page_number, page_data in enumerate(doc):
    if page_number==1:
        txt = pytesseract.image_to_string(page_data).encode("utf-8")
        #print("Page # {} - {}".format(str(page_number),txt))
        #print(txt)
        txt=str(txt)
        txt=txt.split("\\n")
        txt=" ".join(txt)
        #txt.split()
        print(txt.split())
        #print(str(txt))
        #print(str(txt).split())

"""
#print(type(pdf[1]))
#text=pdf[1]

#for e in text.split('\n'):
#    print(len(e),e)
#print(len(text.split('\t')))
#text_file=open("out2.txt","w")
#n=text_file.write(pdf[1])
#text_file.close()
#print(inspect.getmembers(pdf,lambda a:not(inspect.isroutine(a))))
#print(dir(pdftotext.PDF))
#print(type(pdf))
#print(pdfToText(file2)[1])
#pdf=pdfplumber.open(file2)
#print(pdf.pages[1].extract_text())
#test xpdf
#file=__file__
#print('getcwd:      ', os.getcwd())
#print(os.path.dirname(file))
#print(os.sep)


